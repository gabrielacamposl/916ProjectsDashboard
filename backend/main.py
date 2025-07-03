# main.py - Backend completo con nuevas funcionalidades y fechas de remodelación
from flask import Flask, jsonify, send_from_directory
from flask_cors import CORS
import requests
import openpyxl
from io import BytesIO
import schedule
import time
import threading
from datetime import datetime, timedelta
import os
import logging

# Configurar logging para debug
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app)

# Variables globales
dashboard_data = {
    "last_update": None,
    "florida_data": {},
    "texas_data": {},
    "global_data": {},
    "remodel_dates": {},
    "status": "waiting"
}

workbook = None  # Variable global para el workbook

def download_and_process_excel():
    """Descarga el Excel de SharePoint y procesa los datos"""
    global dashboard_data, workbook
    
    try:
        logger.info("Iniciando descarga de SharePoint...")
        
        # URL del SharePoint
        #SharePoint del Excel de remodelación de tiendas Graficas
        url = "https://916foods-my.sharepoint.com/personal/it_support_916foods_com/_layouts/15/download.aspx?share=EZEBqKqQF9pFitMhSuZPwj4B4xV5tW0qtHLdceNN5-I9Ug"
        # Share point del Excel de remodelación de tiendas c/ Marco
        url_shp = "https://916foods-my.sharepoint.com/personal/it_support_916foods_com/_layouts/15/download.aspx?share=EZb5NHihKQ9Lnysp--9gH0UBOkCr7K-3Ud_mPhC2At2PPQ"
        # Headers para evitar bloqueos
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        
        response = requests.get(url, headers=headers, timeout=30)
        logger.info(f"Respuesta SharePoint: Status {response.status_code}, Tamaño: {len(response.content)} bytes")
        
        response.raise_for_status()
        
        # Verificar que el archivo no esté vacío
        if len(response.content) < 1000:
            raise Exception(f"Archivo muy pequeño o vacío: {len(response.content)} bytes")
        
        # Cargar el Excel en memoria
        logger.info("Cargando archivo Excel...")
        workbook = openpyxl.load_workbook(BytesIO(response.content), data_only=True)
        
        logger.info(f"Hojas encontradas en Excel: {workbook.sheetnames}")
        
        # Verificar que existen las hojas necesarias
        required_sheets = ['FLO', 'TEX']
        for sheet_name in required_sheets:
            if sheet_name not in workbook.sheetnames:
                raise Exception(f"Hoja '{sheet_name}' no encontrada. Disponibles: {workbook.sheetnames}")
        
        logger.info("Archivo Excel cargado correctamente")
        
        # Procesar datos de Florida
        logger.info("Procesando datos de Florida...")
        florida_data = process_sheet_data(workbook, 'FLO')
        
        # Procesar datos de Texas  
        logger.info("Procesando datos de Texas...")
        texas_data = process_sheet_data(workbook, 'TEX')
        
        # Combinar datos globales
        logger.info("Combinando datos globales...")
        global_data = combine_regional_data(florida_data, texas_data)
        
        # Obtener fechas de remodelación
        logger.info("Obteniendo fechas de remodelación...")
        remodel_dates = get_remodel_dates()
        
        # Actualizar datos globales
        dashboard_data = {
            "last_update": datetime.now().isoformat(),
            "florida_data": florida_data,
            "texas_data": texas_data,
            "global_data": global_data,
            "remodel_dates": remodel_dates,
            "status": "success"
        }
        
        logger.info("Datos procesados correctamente")
        logger.info(f"Resumen - FL: {florida_data.get('aloha19', {}).get('total', 0)} tiendas, TX: {texas_data.get('aloha19', {}).get('total', 0)} tiendas")
        logger.info(f"Fechas de remodelación: Stage 1: {remodel_dates.get('stage1_start', 'TBD')} → {remodel_dates.get('stage1_end', 'TBD')}")
        
    except Exception as e:
        error_msg = f"Error procesando datos: {str(e)}"
        logger.error(f"{error_msg}")
        dashboard_data["status"] = f"error: {str(e)}"

def read_excel_cell(sheet, cell):
    """Lee una celda del Excel de forma segura con DEBUG"""
    try:
        cell_obj = sheet[cell]
        value = cell_obj.value
        
        # DEBUG: Mostrar valor exacto de cada celda
        logger.info(f"Celda {cell}: '{value}' (tipo: {type(value)})")
        
        if value is None:
            logger.warning(f" Celda {cell} está vacía")
            return 0
        
        # Convertir a número
        if isinstance(value, (int, float)):
            num_value = float(value)
        else:
            # Si es texto, intentar convertir
            try:
                num_value = float(str(value).strip())
            except:
                logger.warning(f"No se pudo convertir '{value}' a número en celda {cell}")
                return 0
        
        # Validar que no sea negativo
        if num_value < 0:
            logger.warning(f"Valor negativo en celda {cell}: {num_value}")
            return 0
        
        logger.info(f"Celda {cell} = {num_value}")
        return num_value
        
    except Exception as e:
        logger.error(f"Error leyendo celda {cell}: {str(e)}")
        return 0

def read_excel_date_cell(sheet, cell):
    """Lee una celda que contiene fecha y la formatea correctamente"""
    try:
        cell_obj = sheet[cell]
        value = cell_obj.value
        
        logger.info(f"Celda {cell}: '{value}' (tipo: {type(value)})")
        
        if value is None:
            logger.warning(f"Celda de fecha {cell} está vacía")
            return "TBD"
        
        # Si es una fecha de Excel (datetime)
        if hasattr(value, 'strftime'):
            formatted_date = value.strftime("%m/%d/%Y")  # CAMBIADO: Formato MM/DD/YYYY
            logger.info(f"Fecha {cell} = {formatted_date}")
            return formatted_date
        
        # Si es texto que parece una fecha
        if isinstance(value, str):
            value = value.strip()
            if value.upper() in ["TBD", "PENDING", "---", ""]:
                return "TBD"
            
            # Si contiene hora (formato YYYY-MM-DD HH:MM:SS), extraer solo la fecha
            if " " in value and ":" in value:
                try:
                    # Separar fecha de hora
                    date_part = value.split(" ")[0]
                    # Intentar parsear como YYYY-MM-DD
                    if "-" in date_part:
                        parts = date_part.split("-")
                        if len(parts) == 3:
                            year, month, day = parts
                            formatted_date = f"{month.zfill(2)}/{day.zfill(2)}/{year}"
                            logger.info(f"Fecha formateada de texto con hora {cell} = {formatted_date}")
                            return formatted_date
                except Exception as e:
                    logger.warning(f"Error parseando fecha con hora: {e}")
            
            # Si ya está en formato MM/DD/YYYY o similar
            if "/" in value:
                logger.info(f"Fecha texto {cell} = {value}")
                return value
            # Si está en formato YYYY-MM-DD
            elif "-" in value:
                try:
                    parts = value.split("-")
                    if len(parts) == 3:
                        year, month, day = parts
                        formatted_date = f"{month.zfill(2)}/{day.zfill(2)}/{year}"
                        logger.info(f"Fecha convertida de YYYY-MM-DD {cell} = {formatted_date}")
                        return formatted_date
                except:
                    pass
        
        # Si es un número (días desde 1900)
        if isinstance(value, (int, float)):
            try:
                # Excel epoch: 1 de enero de 1900 (con ajuste por bug de Excel)
                excel_epoch = datetime(1900, 1, 1)
                if value > 0:
                    date_obj = excel_epoch + timedelta(days=value - 2)  # -2 por bug histórico de Excel
                    formatted_date = date_obj.strftime("%m/%d/%Y")  # CAMBIADO: Formato MM/DD/YYYY
                    logger.info(f"Fecha numérica {cell} = {formatted_date}")
                    return formatted_date
            except:
                pass
        
        # Fallback: convertir a string y limpiar
        fallback_value = str(value) if value else "TBD"
        # Si el fallback contiene hora, quitarla
        if " " in fallback_value and ":" in fallback_value:
            fallback_value = fallback_value.split(" ")[0]
        
        logger.warning(f"Formato de fecha no reconocido en {cell}: {value}, usando fallback: {fallback_value}")
        return fallback_value
        
    except Exception as e:
        logger.error(f"Error leyendo fecha en celda {cell}: {str(e)}")
        return "TBD"

def combine_dates(date1, date2):
    """Combina dos fechas, priorizando la que no sea TBD"""
    if date1 and date1 != "TBD":
        return date1
    elif date2 and date2 != "TBD":
        return date2
    else:
        return "TBD"

def get_remodel_dates():
    """Obtiene las fechas de remodelación desde SharePoint (celdas específicas)"""
    try:
        global workbook
        if not workbook:
            logger.warning("No hay workbook disponible para fechas de remodelación")
            return {
                "stage1_start": "TBD",
                "stage1_end": "TBD", 
                "stage2_start": "TBD",
                "stage2_end": "TBD",
                "source": "fallback"
            }
        
        logger.info("=== OBTENIENDO FECHAS DE REMODELACIÓN ===")
        
        # Verificar que existen las hojas necesarias
        required_sheets = ['FLO', 'TEX']
        for sheet_name in required_sheets:
            if sheet_name not in workbook.sheetnames:
                logger.warning(f"Hoja {sheet_name} no encontrada para fechas")
                return {
                    "stage1_start": "TBD",
                    "stage1_end": "TBD",
                    "stage2_start": "TBD", 
                    "stage2_end": "TBD",
                    "source": "fallback - missing sheets"
                }
        
        # Leer fechas de Florida (FLO)
        flo_sheet = workbook['FLO']
        logger.info("Leyendo fechas de Florida...")
        
        flo_stage1_start = read_excel_date_cell(flo_sheet, 'C3')  # Stage 1 Start
        flo_stage1_end = read_excel_date_cell(flo_sheet, 'D3')    # Stage 1 End  
        flo_stage2_start = read_excel_date_cell(flo_sheet, 'C4')  # Stage 2 Start
        flo_stage2_end = read_excel_date_cell(flo_sheet, 'D4')    # Stage 2 End
        
        # NUEVO: Leer cantidad de tiendas por mes en Florida
        flo_july_stores = read_excel_cell(flo_sheet, 'E3')        # Tiendas en Julio
        flo_august_stores = read_excel_cell(flo_sheet, 'F3')      # Tiendas en Agosto
        
        logger.info(f"Florida - Julio: {flo_july_stores} tiendas, Agosto: {flo_august_stores} tiendas")
        
        # Leer fechas de Texas (TEX) - CORREGIDO según especificación del usuario
        tex_sheet = workbook['TEX']
        logger.info("Leyendo fechas de Texas...")
        
        tex_stage1_start = read_excel_date_cell(tex_sheet, 'C3')  # Stage 1 Start Remod
        tex_stage1_end = read_excel_date_cell(tex_sheet, 'D3')    # Stage 1 End Remod  
        tex_stage2_start = read_excel_date_cell(tex_sheet, 'C4')  # Stage 2 Start Remod
        tex_stage2_end = read_excel_date_cell(tex_sheet, 'D4')    # Stage 2 End Remod
        
        # NUEVO: Leer cantidad de tiendas por mes en Texas
        tex_june_stores = read_excel_cell(tex_sheet, 'E3')        # Tiendas en Junio
        tex_july_stores = read_excel_cell(tex_sheet, 'F3')       # Tiendas en Julio
        
        logger.info(f"Texas - Junio: {tex_june_stores} tiendas, Julio: {tex_july_stores} tiendas")
        
        # Combinar fechas (usar la primera válida encontrada o la más temprana)
        stage1_start = combine_dates(flo_stage1_start, tex_stage1_start)
        stage1_end = combine_dates(flo_stage1_end, tex_stage1_end)
        stage2_start = combine_dates(flo_stage2_start, tex_stage2_start)
        stage2_end = combine_dates(flo_stage2_end, tex_stage2_end)
        
        result = {
            "stage1_start": stage1_start,
            "stage1_end": stage1_end,
            "stage2_start": stage2_start,
            "stage2_end": stage2_end,
            "source": "sharepoint",
            "regional_details": {
                "florida": {
                    "stage1_start": flo_stage1_start,
                    "stage1_end": flo_stage1_end,
                    "stage2_start": flo_stage2_start,
                    "stage2_end": flo_stage2_end,
                    "july_stores": flo_july_stores,      
                    "august_stores": flo_august_stores   
                },
                "texas": {
                    "stage1_start": tex_stage1_start,
                    "stage1_end": tex_stage1_end,
                    "stage2_start": tex_stage2_start,
                    "stage2_end": tex_stage2_end,
                    "june_stores": tex_june_stores,     
                    "july_stores": tex_july_stores       
                }
            }
        }
        
        return result
        
    except Exception as e:
        logger.error(f"Error obteniendo fechas de remodelación: {str(e)}")
        return {
            "stage1_start": "TBD",
            "stage1_end": "TBD",
            "stage2_start": "TBD",
            "stage2_end": "TBD",
            "source": f"error: {str(e)}"
        }

def process_sheet_data(workbook, sheet_name):
    """Procesa los datos de una hoja específica (FLO o TEX) con DEBUG completo"""
    try:
        sheet = workbook[sheet_name]
        logger.info(f"Procesando hoja: {sheet_name}")
        
        # DEBUG: Mostrar información básica de la hoja
        logger.info(f"Dimensiones de la hoja: {sheet.max_row} filas x {sheet.max_column} columnas")
        
        if sheet_name == 'FLO':
            logger.info(" === PROCESANDO FLORIDA (FLO) ===")
            
            # DEBUG: Leer y mostrar cada celda individual
            logger.info("Leyendo datos de Aloha 19...")
            # stage1 = read_excel_cell(sheet, 'B3')
            # Tiendas que se van a remodelar en Florida por meses
            # Julio y Agosto
            july = read_excel_cell(sheet, 'E3')  # Total de tiendas en junio
            august = read_excel_cell(sheet, 'F3')  # Total de tiendas en agosto
            stage2 = read_excel_cell(sheet, 'B4')
            finished = read_excel_cell(sheet, 'B5')
            total = read_excel_cell(sheet, 'B6')
            
            # CORREGIDO: Florida wiring debe ser B11 (finished) y B12 (pending)
            wiring_finished = read_excel_cell(sheet, 'B11')  # Cambiado de B10 a B11
            wiring_pending = read_excel_cell(sheet, 'B12')   # Cambiado de B11 a B12
            
            logger.info("Leyendo datos de Tecnologías (Columna C - YES)...")
            fresh_ai = read_excel_cell(sheet, 'C15')
            edmb = read_excel_cell(sheet, 'C16')
            idmb = read_excel_cell(sheet, 'C17')
            qb = read_excel_cell(sheet, 'C18')
            kiosk = read_excel_cell(sheet, 'C19')
            
            logger.info("Leyendo datos de Proyectos...")
            signed = read_excel_cell(sheet, 'B24')
            quote = read_excel_cell(sheet, 'B25')
            paid = read_excel_cell(sheet, 'B26')
            
            logger.info("Leyendo tipos de proyectos...")
            project_edmb_idmb_qb = read_excel_cell(sheet, 'B30')
            project_fai_edmb_idmb_qb = read_excel_cell(sheet, 'B31')
            
            # Datos de Florida
            data = {
                "aloha19": {
                    #"stage1": stage1,
                    "july": july,  # Total de tiendas en junio
                    "august": august,  # Total de tiendas en agosto
                    "stage2": stage2, 
                    "finished": finished,
                    "total": total
                },
                "wiring": {
                    "pending": wiring_pending,   
                    "finished": wiring_finished  
                },
                "technologies": {
                    "fresh_ai": fresh_ai,
                    "edmb": edmb,
                    "idmb": idmb,
                    "qb": qb,
                    "kiosk": kiosk
                },
                "projects": {
                    "signed": signed,
                    "quote": quote,
                    "paid": paid
                },
                "project_types": {
                    "edmb_idmb_qb": project_edmb_idmb_qb,
                    "fai_edmb_idmb_qb": project_fai_edmb_idmb_qb
                }
            }
            
        else:  # TEX
            logger.info("=== PROCESANDO TEXAS (TEX) ===")
            
            logger.info("Leyendo datos de Aloha 19...")
           # stage1 = read_excel_cell(sheet, 'B3')
            # Tiendas que se van a remodelar en Texas por meses
            june = read_excel_cell(sheet, 'E3')  # Total de tiendas en junio
            july = read_excel_cell(sheet, 'F3')  # Total de tiendas en julio
            stage2 = read_excel_cell(sheet, 'B4')
            close = read_excel_cell(sheet, 'B5')
            finished = read_excel_cell(sheet, 'B6')
            total = read_excel_cell(sheet, 'B7')
            
            logger.info("🔌 Leyendo datos de Wiring...")
            wiring_pending = read_excel_cell(sheet, 'B12')
            wiring_finished = read_excel_cell(sheet, 'B13')
            wiring_close = read_excel_cell(sheet, 'B14')  # NUEVO: Wiring Close (CELDA B14)
            
            logger.info(f"Texas Wiring - Pending: {wiring_pending}, Finished: {wiring_finished}, Close: {wiring_close}")
            
            # VALIDAR que Close no sea 0 para enviarlo al frontend
            if wiring_close and wiring_close > 0:
                logger.info(f"Texas Wiring Close válido: {wiring_close}")
            else:
                logger.warning(f"Texas Wiring Close es 0 o inválido: {wiring_close}")
            
            fresh_ai = read_excel_cell(sheet, 'B18')
            edmb = read_excel_cell(sheet, 'B19')
            idmb = read_excel_cell(sheet, 'B20')
            qb = read_excel_cell(sheet, 'B21')
            kiosk = read_excel_cell(sheet, 'B22')

            paid = read_excel_cell(sheet, 'B26')  
            signed = read_excel_cell(sheet, 'B27') 
            quote = read_excel_cell(sheet, 'B28')  # TEXAS Quote está en B28
        
            project_edmb = read_excel_cell(sheet, 'B33')
            project_edmb_idmb_qb = read_excel_cell(sheet, 'B34')  # TEX con EDMB-IDMB-QB
            project_idmb_qb = read_excel_cell(sheet, 'B35')  # TEX con IDMB y QB
            
            # Datos de Texas
            data = {
                "aloha19": {
                    #"stage1": stage1,
                    "june": june,  # Total de tiendas en junio
                    "july": july,  # Total de tiendas en julio
                    "stage2": stage2,
                    "close": close,
                    "finished": finished,
                    "total": total
                },
                "wiring": {
                    "pending": wiring_pending,
                    "finished": wiring_finished,
                    "close": wiring_close 
                },
                "technologies": {
                    "fresh_ai": fresh_ai,
                    "edmb": edmb,
                    "idmb": idmb,
                    "qb": qb,
                    "kiosk": kiosk
                },
                "projects": {
                    "paid": paid,  # Texas Quote = 5 (B27)
                    "signed": signed,  # Texas Pending (B28)
                    "quote": quote,  # Texas Quote
                },
                "project_types": {
                    "edmb": project_edmb,
                    "edmb_idmb_qb": project_edmb_idmb_qb,  # TEX con EDMB-IDMB-QB 
                    "idmb_qb": project_idmb_qb  # TEX con IDMB y QB
                    
                }
            }
        
        logger.info(f"Datos procesados para {sheet_name}:")
        logger.info(f"   Aloha19 Total: {data['aloha19']['total']}")
        logger.info(f"   Aloha19 Finished: {data['aloha19']['finished']}")
        logger.info(f"   Wiring Finished: {data['wiring']['finished']}")
        logger.info(f"   Wiring Pending: {data['wiring']['pending']}")
        if sheet_name == 'TEX':
            logger.info(f"Wiring Close: {data['wiring'].get('close', 0)}")
        logger.info(f"Fresh AI: {data['technologies']['fresh_ai']}")
        
        return data
        
    except Exception as e:
        logger.error(f"Error procesando hoja {sheet_name}: {str(e)}")
        return {}

def combine_regional_data(florida_data, texas_data):
    """Combina los datos de Florida y Texas para vista global con DEBUG"""
    try:
        logger.info(" === COMBINANDO DATOS GLOBALES ===")
        
        # Helper para obtener valores seguros
        def safe_get(data, path, default=0):
            try:
                result = data
                for key in path.split('.'):
                    result = result[key]
                return result or default
            except:
                return default
        
        fl_total = safe_get(florida_data, 'aloha19.total')
        tx_total = safe_get(texas_data, 'aloha19.total')
        
        fl_finished = safe_get(florida_data, 'aloha19.finished')
        tx_finished = safe_get(texas_data, 'aloha19.finished')
        
        logger.info(f" Florida - Total: {fl_total}, Finished: {fl_finished}")
        logger.info(f"Texas - Total: {tx_total}, Finished: {tx_finished}")
        
        # CORREGIDO: Combinar proyectos de ambas regiones
        fl_signed = safe_get(florida_data, 'projects.signed')
        fl_quote = safe_get(florida_data, 'projects.quote')
        fl_paid = safe_get(florida_data, 'projects.paid')
        
        tx_paid = safe_get(texas_data, 'projects.paid')
        tx_signed = safe_get(texas_data, 'projects.signed')
        tx_quote = safe_get(texas_data, 'projects.quote')
        tx_wiring_close = safe_get(texas_data, 'wiring.close')  # NUEVO: Wiring Close de Texas
        
        logger.info(f"Florida Projects - Signed: {fl_signed}, Quote: {fl_quote}, Paid: {fl_paid}")
        logger.info(f"Texas Projects - Quote: {tx_quote}, Signed: {tx_signed}, Paid: {tx_paid}")
        logger.info(f"Texas Wiring Close: {tx_wiring_close}")
        
        global_data = {
            "aloha19": {
               # "stage1": safe_get(florida_data, 'aloha19.july') + safe_get(texas_data, 'aloha19.stage1'),
                "june":  safe_get(texas_data, 'aloha19.june') ,  # Solo Texas tiene junio
                "july": safe_get(florida_data, 'aloha19.august') + safe_get(texas_data, 'aloha19.july'),
                "august": safe_get(florida_data, 'aloha19.august'),  # Solo Florida tiene agosto
                "stage2": safe_get(florida_data, 'aloha19.stage2') + safe_get(texas_data, 'aloha19.stage2'),
                "close": safe_get(texas_data, 'aloha19.close'),  # Solo Texas tiene "close"
                "finished": fl_finished + tx_finished,
                "total": fl_total + tx_total
            },
            "wiring": {
                "pending": safe_get(florida_data, 'wiring.pending') + safe_get(texas_data, 'wiring.pending'),
                "finished": safe_get(florida_data, 'wiring.finished') + safe_get(texas_data, 'wiring.finished'),
                "close": safe_get(texas_data, 'wiring.close')  # NUEVO: Solo Texas tiene wiring close
            },
            "technologies": {
                "fresh_ai": safe_get(florida_data, 'technologies.fresh_ai') + safe_get(texas_data, 'technologies.fresh_ai'),
                "edmb": safe_get(florida_data, 'technologies.edmb') + safe_get(texas_data, 'technologies.edmb'),
                "idmb": safe_get(florida_data, 'technologies.idmb') + safe_get(texas_data, 'technologies.idmb'),
                "qb": safe_get(florida_data, 'technologies.qb') + safe_get(texas_data, 'technologies.qb'),
                "kiosk": safe_get(florida_data, 'technologies.kiosk') + safe_get(texas_data, 'technologies.kiosk')
            },
            # CORREGIDO: Combinar proyectos globales
            "projects": {
                "signed": fl_signed + tx_signed,  # Solo Florida tiene signed
                "quote": fl_quote + tx_quote,  # Florida + Texas quotes
                "paid": fl_paid + tx_paid,  # Solo Florida tiene paid
                
            },
            # AGREGADO: Datos separados para gráficas individuales
            "project_types_florida": {
                "edmb_idmb_qb": safe_get(florida_data, 'project_types.edmb_idmb_qb'),
                "fai_edmb_idmb_qb": safe_get(florida_data, 'project_types.fai_edmb_idmb_qb')
            },
            "project_types_texas": {
                "edmb": safe_get(texas_data, 'project_types.edmb'),
                "edmb_idmb": safe_get(texas_data, 'project_types.edmb_idmb'),
                "idmb": safe_get(texas_data, 'project_types.idmb')
            }
        }
        
        logger.info(f"Global combinado - Total: {global_data['aloha19']['total']}, Finished: {global_data['aloha19']['finished']}")
        logger.info(f"Global Fresh AI: {global_data['technologies']['fresh_ai']}")
        
        return global_data
        
    except Exception as e:
        logger.error(f"Error combinando datos: {str(e)}")
        return {}

# FUNCIONES PARA TABLAS DETALLADAS

def get_table_data(sheet_name, columns=None, filter_rows=True, max_row=None):
    """Obtiene datos de una hoja para tabla con filtros opcionales"""
    try:
        global workbook
        if not workbook or sheet_name not in workbook.sheetnames:
            return {"error": f"Hoja {sheet_name} no encontrada"}
        
        sheet = workbook[sheet_name]
        logger.info(f"Leyendo tabla de hoja: {sheet_name}")
        
        # Definir rangos específicos por hoja
        if sheet_name == 'TEX-COM':
            actual_max_row = min(59, sheet.max_row) if max_row is None else max_row
        elif sheet_name == 'FLO-COM':
            actual_max_row = min(28, sheet.max_row) if max_row is None else max_row
        else:
            actual_max_row = sheet.max_row if max_row is None else max_row
        
        # Si no se especifican columnas, leer de A hasta W (23)
        if not columns:
            columns = [chr(65 + i) for i in range(24)]  # A-x
        
        table_data = []
        
        # Mapeo de columnas a nombres legibles (basado en el diagnóstico real)
        column_names = {
            'A': 'STORE', 'B': 'ADDRESS', 'C': 'PHONE/STORE PHONE', 'D': 'DM', 'E': 'GM',
            'F': 'A19', 'G': 'WIRING', 'H': 'FRESH AI', 'I': 'EDMB', 'J': 'IDMB',
            'K': 'QB', 'L': 'KIOSK', 'M': 'A19 UP', 'N':'NETXEO PRO', 'O': 'START REMOD', 'P': 'END REMOD',
            'Q': 'PROJECT', 'R': 'AUV', 'S': 'COST', 'T': 'STATUS', 'U': 'CABLE INSTALL',
            'V': 'DELIVERY DATE', 'W': 'INSTALLATION DATE', 'X': 'INSTALL'
        }
        
        # Columnas que contienen fechas (no convertir 0 a "---")
        date_columns = ['M', 'N', 'O', 'P', 'U', 'V', 'W']
        
        # Leer datos fila por fila (empezar desde fila 2 para evitar headers)
        for row_num in range(2, actual_max_row + 1):
            row_data = {}
            valid_row = False
            
            for col in columns:
                try:
                    cell_value = sheet[f"{col}{row_num}"].value
                    # Manejo especial para columnas de fecha
                    if col in date_columns:
                        if cell_value is None:
                            cell_value = "---"
                        else:
                            # Para fechas, usar la función de formateo de fecha
                            formatted_date = read_excel_date_cell(sheet, f"{col}{row_num}")
                            cell_value = formatted_date if formatted_date != "TBD" else "---"
                    else:
                        # Para otras columnas, manejo normal
                        if cell_value is None:
                            cell_value = "---"
                        elif isinstance(cell_value, (int, float)) and cell_value == 0:
                            cell_value = "---"  # Cambiar 0 por "---" solo en columnas no-fecha
                        else:
                            cell_value = str(cell_value).strip()
                            if cell_value in ["", "0", "0.0"]:
                                cell_value = "---"
                    
                    row_data[col] = cell_value
                    row_data[f"{col}_name"] = column_names.get(col, f"Col_{col}")
                    
                    # Marcar como fila válida si tiene contenido real
                    if cell_value not in ["---", "", " "]:
                        valid_row = True
                        
                except Exception as e:
                    logger.error(f"Error leyendo celda {col}{row_num}: {str(e)}")
                    row_data[col] = "---"
                    row_data[f"{col}_name"] = column_names.get(col, f"Col_{col}")
            
            # Agregar fila solo si es válida o si no estamos filtrando
            if valid_row or not filter_rows:
                table_data.append({"row": row_num, "data": row_data})
        
        logger.info(f"Tabla {sheet_name} leída: {len(table_data)} filas (rango hasta fila {actual_max_row})")
        
        return {"data": table_data, "columns": columns, "column_names": column_names}
        
    except Exception as e:
        logger.error(f"Error leyendo tabla {sheet_name}: {str(e)}")
        return {"error": str(e)}

@app.route('/')
def home():
    return jsonify({
        "message": "916 Foods Dashboard API",
        "status": dashboard_data["status"],
        "last_update": dashboard_data["last_update"]
    })

@app.route('/api/data')
def get_dashboard_data():
    """Endpoint principal que devuelve todos los datos"""
    logger.info(f"API request - Status: {dashboard_data['status']}")
    return jsonify(dashboard_data)

@app.route('/api/florida')
def get_florida_data():
    """Endpoint para datos solo de Florida"""
    return jsonify({
        "data": dashboard_data["florida_data"],
        "last_update": dashboard_data["last_update"],
        "status": dashboard_data["status"]
    })

@app.route('/api/texas') 
def get_texas_data():
    """Endpoint para datos solo de Texas"""
    return jsonify({
        "data": dashboard_data["texas_data"],
        "last_update": dashboard_data["last_update"],
        "status": dashboard_data["status"]
    })

@app.route('/api/refresh')
def manual_refresh():
    """Endpoint para forzar actualización manual"""
    logger.info("Refresh manual solicitado")
    threading.Thread(target=download_and_process_excel).start()
    return jsonify({"message": "Actualización iniciada"})

@app.route('/api/remodel-dates')
def get_remodel_dates_api():
    """Endpoint para obtener fechas de remodelación desde SharePoint"""
    try:
        logger.info("API request - Fechas de remodelación")
        dates = dashboard_data.get("remodel_dates", {})
        
        # Si no hay fechas en dashboard_data, intentar obtenerlas directamente
        if not dates or dates.get("source") == "fallback":
            dates = get_remodel_dates()
        
        return jsonify({
            "status": "success",
            "last_update": dashboard_data.get("last_update"),
            **dates
        })
        
    except Exception as e:
        logger.error(f"Error en endpoint fechas de remodelación: {str(e)}")
        return jsonify({
            "status": "error",
            "stage1_start": "TBD",
            "stage1_end": "TBD", 
            "stage2_start": "TBD",
            "stage2_end": "TBD",
            "source": f"error: {str(e)}"
        })

# ENDPOINTS PARA TABLAS DETALLADAS
@app.route('/api/table/<region>/detailed')
def get_detailed_regional_table(region):
    """Obtiene tabla detallada regional de hojas FLO-COM o TEX-COM"""
    try:
        if region.lower() == 'florida':
            sheet_name = 'FLO-COM'
        elif region.lower() == 'texas':
            sheet_name = 'TEX-COM'
        else:
            return jsonify({"error": "Región debe ser 'florida' o 'texas'"})
        
        # Leer toda la tabla de la hoja COM
        result = get_table_data(sheet_name, filter_rows=True)
        
        if "error" in result:
            return jsonify(result)
        
        return jsonify({
            "status": "success",
            "region": region,
            "sheet": sheet_name,
            "data": result["data"],
            "columns": result["columns"],
            "total_rows": len(result["data"])
        })
        
    except Exception as e:
        logger.error(f"Error en tabla detallada {region}: {str(e)}")
        return jsonify({"error": str(e)})

@app.route('/api/table/projects')
def get_project_details_table():
    """Obtiene tabla de detalles de proyectos con columnas específicas y filtros"""
    try:
        # Columnas específicas requeridas según diagnóstico:
        # A=STORE, B=ADDRESS, M=A19 UP, 
        required_columns = ['A', 'B', 'M', 'N', 'Q', 'R', 'S', 'T', 'U', 'V', 'W']
        
        # Filtros válidos para la columna PROJECT   
        valid_projects = ['FAI,EDMB,IDMB,QUE', 'EDMB-IDMB-QB', 'EDMB', 'EDMB-IDMB-QB', 'IDMB-QB']
        
        # Intentar con ambas hojas
        project_data = []
        debug_info = {
            "sheets_processed": [],
            "total_rows_checked": 0,
            "rows_with_project_data": 0,
            "rows_matching_filters": 0
        }
        
        for sheet_name in ['FLO-COM', 'TEX-COM']:
            if sheet_name not in [s for s in (workbook.sheetnames if workbook else [])]:
                logger.warning(f"Hoja {sheet_name} no encontrada")
                debug_info["sheets_processed"].append(f"{sheet_name}: NO_EXISTE")
                continue
                
            result = get_table_data(sheet_name, required_columns, filter_rows=False)
            
            if "error" not in result:
                debug_info["sheets_processed"].append(f"{sheet_name}: PROCESADA")
                
                # Filtrar filas según PROJECT 
                for row_info in result["data"]:
                    row_data = row_info["data"]
                    debug_info["total_rows_checked"] += 1
                    
                    # ¿ PROJECT está en columna P
                    project_value = row_data.get('Q', '---').strip()
                    
                    if project_value not in ['---', '', ' ', 'NULL', '-----']:
                        debug_info["rows_with_project_data"] += 1
                        logger.info(f"Proyecto encontrado en {sheet_name} fila {row_info['row']}: '{project_value}'")
                    
                    # Verificar si el proyecto coincide con alguno de los filtros válidos
                    project_matches = False
                    for valid_project in valid_projects:
                        if valid_project.upper() in project_value.upper():
                            project_matches = True
                            debug_info["rows_matching_filters"] += 1
                            logger.info(f"Coincidencia encontrada: '{project_value}' contiene '{valid_project}'")
                            break
                    
                    # Solo incluir si tiene un proyecto válido y datos completos
                    if project_matches:
                        # Verificar que tenga al menos STORE o ADDRESS válidos
                        store = row_data.get('A', '---')
                        address = row_data.get('B', '---')
                        
                        if store not in ['---', '', ' '] or address not in ['---', '', ' ']:
                            # Agregar información de la hoja de origen
                            row_data['_source_sheet'] = sheet_name
                            row_data['_region'] = 'Florida' if sheet_name == 'FLO-COM' else 'Texas'
                            project_data.append(row_info)
                            logger.info(f"Fila válida agregada: Store={store}, Project={project_value}")
            else:
                debug_info["sheets_processed"].append(f"{sheet_name}: ERROR - {result['error']}")
                logger.error(f"Error procesando {sheet_name}: {result['error']}")
        
        # Log de resumen
        logger.info(f"RESUMEN: {debug_info['total_rows_checked']} filas revisadas, {debug_info['rows_with_project_data']} con datos de proyecto, {debug_info['rows_matching_filters']} coinciden con filtros")
        
        # Mapeo de nombres de columnas según diagnóstico
        column_display_names = {
            'A': 'STORE',
            'B': 'ADDRESS', 
            'M': 'A19 UP',   
            'N': 'NETXEO PRO',
            'Q': 'PROJECT',  
            'R': 'AUV',      
            'S': 'COST',     
            'T': 'STATUS',
            'U': 'CABLE INSTALL',   
            'V': 'DELIVERY DATE',
            'W': 'INSTALLATION DATE',
            'X': 'INSTALL'   
        }
        
        return jsonify({
            "status": "success",
            "data": project_data,
            "columns": required_columns,
            "column_names": column_display_names,
            "total_rows": len(project_data),
            "debug_info": debug_info,
            "filters_applied": {
                "project_types": valid_projects,
                "note": "Solo se muestran filas con PROJECT"
            }
        })
        
    except Exception as e:
        logger.error(f"Error en tabla de proyectos: {str(e)}")
        return jsonify({"error": str(e)})

# ENDPOINTS DE DEBUG Y UTILIDAD

@app.route('/api/debug')
def debug_info():
    """Endpoint para información de debug"""
    return jsonify({
        "status": dashboard_data["status"],
        "last_update": dashboard_data["last_update"],
        "remodel_dates_status": dashboard_data.get("remodel_dates", {}).get("source", "not_loaded"),
        "data_summary": {
            "florida_total": dashboard_data.get("florida_data", {}).get("aloha19", {}).get("total", 0),
            "texas_total": dashboard_data.get("texas_data", {}).get("aloha19", {}).get("total", 0),
            "global_total": dashboard_data.get("global_data", {}).get("aloha19", {}).get("total", 0),
            "florida_wiring_pending": dashboard_data.get("florida_data", {}).get("wiring", {}).get("pending", 0),
            "florida_wiring_finished": dashboard_data.get("florida_data", {}).get("wiring", {}).get("finished", 0),
            "texas_wiring_close": dashboard_data.get("texas_data", {}).get("wiring", {}).get("close", 0)
        }
    })

@app.route('/api/sheets-available')
def list_available_sheets():
    """Lista todas las hojas disponibles en el Excel"""
    try:
        global workbook
        if not workbook:
            return jsonify({"error": "No workbook loaded"})
        
        return jsonify({
            "status": "success",
            "sheets": workbook.sheetnames,
            "required_for_dashboard": ["FLO", "TEX"],
            "required_for_tables": ["FLO-COM", "TEX-COM"],
            "com_sheets_available": {
                "FLO-COM": "FLO-COM" in workbook.sheetnames,
                "TEX-COM": "TEX-COM" in workbook.sheetnames
            }
        })
        
    except Exception as e:
        return jsonify({"error": str(e)})


def run_scheduler():
    """Ejecuta el scheduler en un hilo separado"""
    while True:
        schedule.run_pending()
        time.sleep(60)

if __name__ == '__main__':
    logger.info("Iniciando 916 Foods Dashboard API...")
    
    # Configurar actualizaciones automáticas cada 30 minutos
    schedule.every(30).minutes.do(download_and_process_excel)
    
    # Ejecutar una vez al inicio
    logger.info("Carga inicial de datos...")
    download_and_process_excel()
    
    # Iniciar scheduler en hilo separado
    scheduler_thread = threading.Thread(target=run_scheduler)
    scheduler_thread.daemon = True
    scheduler_thread.start()
    
    # Iniciar servidor Flask
    port = int(os.environ.get('PORT', 5000))
    logger.info(f"Servidor iniciando en puerto {port}")
    app.run(host='0.0.0.0', port=port, debug=False)